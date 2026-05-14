"""Agent 7 — Monitor.

Weekly AI visibility tracker. Re-queries the 14 audit prompts across ChatGPT,
Perplexity, Claude, Gemini and writes results back into the Flo_GEO_Prompt_Tracking
workbook + a JSON snapshot.

In production this calls each AI engine's API directly:
  - ChatGPT: OpenAI API with browsing tools (or via tryprofound.com / Otterly)
  - Perplexity: api.perplexity.ai
  - Claude: Anthropic API with web search tool
  - Gemini: Google AI Studio API with grounding

This implementation uses the Anthropic web_search tool as the canonical engine
and a single LLM-judge call to score each result. Other engines plug in by
adding an `_engine_query_*` function with the same signature.
"""
import json
import os
from datetime import datetime
from pathlib import Path

import anthropic
from openpyxl import load_workbook

from ._common import OUTPUT, CONFIG, ROOT, load_json, save_json, call_model, cfg, client


JUDGE_SYSTEM = """You are scoring whether and how the brand "Flo Health" is represented in an AI search response.

Given a prompt, an AI engine's answer, and the brand name, output strict JSON:

{
  "cited": true/false,
  "position": "1" | "2" | "3+" | "—" | "Absent",
  "sentiment": 1-5,
  "specificity": "Generic" | "Specific" | "Cited URL",
  "competitors_named": [...],
  "key_quote": "...",
  "source_urls": [...],
  "notes": "..."
}

Rubric:
- Position: first-named app/product wins; if not an apps list, use "—".
- Sentiment: 1=anti-rec, 2=balanced-negative, 3=neutral, 4=positive, 5=strongly positive.
- Specificity: Generic (brand name only), Specific (with a verifiable claim), Cited URL (flo.health URL in citations).
"""


def query_engine_claude(prompt: str) -> str:
    """Use Claude with web search tool as the engine."""
    c = client()
    msg = c.messages.create(
        model=cfg()["model"]["primary"],
        max_tokens=2000,
        tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 5}],
        messages=[{"role": "user", "content": prompt}],
    )
    return "".join(b.text for b in msg.content if b.type == "text")


def query_engine_chatgpt(prompt: str) -> str:
    """Placeholder — wire OpenAI API with browsing here, or call via Profound/Otterly."""
    return f"[chatgpt-stub: enable via OPENAI_API_KEY and add query handler. Prompt: {prompt[:80]}...]"


def query_engine_perplexity(prompt: str) -> str:
    """Placeholder — wire Perplexity API here (api.perplexity.ai)."""
    return f"[perplexity-stub: enable via PERPLEXITY_API_KEY. Prompt: {prompt[:80]}...]"


def query_engine_gemini(prompt: str) -> str:
    """Placeholder — wire Google AI Studio Gemini API with grounding here."""
    return f"[gemini-stub: enable via GOOGLE_API_KEY. Prompt: {prompt[:80]}...]"


ENGINES = {
    "claude": query_engine_claude,
    "chatgpt": query_engine_chatgpt,
    "perplexity": query_engine_perplexity,
    "gemini": query_engine_gemini,
}


def judge(prompt: str, answer: str, brand: str = "Flo Health") -> dict:
    user = (
        f"Prompt: {prompt}\n\n"
        f"AI engine answer:\n{answer[:3000]}\n\n"
        f"Brand to score: {brand}\n\n"
        "Score per the system rubric. JSON only."
    )
    raw = call_model(system=JUDGE_SYSTEM, user=user, max_tokens=1000,
                     temperature=0, json_mode=True)
    return json.loads(raw.strip("` \n").replace("```json", "").replace("```", ""))


def append_to_workbook(results: list[dict], workbook_path: Path) -> None:
    """Append rows into Monthly Tracking sheet of the GEO workbook."""
    if not workbook_path.exists():
        print(f"  [7/7] workbook not found at {workbook_path}, skipping write-back")
        return
    wb = load_workbook(workbook_path)
    ws = wb["Monthly Tracking"]
    next_row = ws.max_row + 1
    for r in results:
        ws.cell(row=next_row, column=1, value=r["month"])
        ws.cell(row=next_row, column=2, value=r["prompt_id"])
        # cols 3-4 are VLOOKUP formulas already in the template
        ws.cell(row=next_row, column=5, value=r["engine"])
        ws.cell(row=next_row, column=6, value="Y" if r["cited"] else "N")
        ws.cell(row=next_row, column=7, value=r["position"])
        ws.cell(row=next_row, column=8, value=r["sentiment"] if r["cited"] else None)
        ws.cell(row=next_row, column=9, value=r["specificity"])
        ws.cell(row=next_row, column=10, value=", ".join(r["competitors_named"]))
        ws.cell(row=next_row, column=11, value=r["key_quote"][:300])
        ws.cell(row=next_row, column=12, value="; ".join(r["source_urls"][:5]))
        ws.cell(row=next_row, column=13, value=r["notes"])
        next_row += 1
    wb.save(workbook_path)
    print(f"  [7/7] appended {len(results)} rows to {workbook_path}")


def run(engines: list[str] | None = None) -> dict:
    audit = load_json(CONFIG / "audit_input.json")
    prompts = audit["prompts"]
    engines = engines or cfg()["monitoring"]["engines"]
    month = datetime.now().strftime("%Y-%m")

    results: list[dict] = []
    for p in prompts:
        for engine in engines:
            handler = ENGINES.get(engine)
            if not handler:
                continue
            try:
                answer = handler(p["text"])
                score = judge(p["text"], answer)
                row = {
                    "month": month,
                    "prompt_id": p["id"],
                    "engine": engine.title() if engine != "chatgpt" else "ChatGPT",
                    "cited": score.get("cited", False),
                    "position": score.get("position", "Absent"),
                    "sentiment": score.get("sentiment", 0),
                    "specificity": score.get("specificity", "Generic"),
                    "competitors_named": score.get("competitors_named", []),
                    "key_quote": score.get("key_quote", ""),
                    "source_urls": score.get("source_urls", []),
                    "notes": score.get("notes", ""),
                }
                results.append(row)
                print(f"  [7/7] {p['id']} × {engine}: cited={row['cited']} sent={row['sentiment']}")
            except Exception as e:
                print(f"  [7/7] FAILED {p['id']} × {engine}: {e}")

    out = OUTPUT / f"07_monitor_{month}.json"
    save_json(out, {"results": results, "month": month, "engines": engines})

    workbook = ROOT.parent / "Flo_GEO_Prompt_Tracking.xlsx"
    append_to_workbook(results, workbook)
    print(f"[7/7] Monitor → {len(results)} observations → {out}")
    return {"results": results}


if __name__ == "__main__":
    run()
